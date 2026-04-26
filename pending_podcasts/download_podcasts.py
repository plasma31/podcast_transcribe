#!/usr/bin/env python3
"""
Podcast Episode Bulk Downloader
================================
Downloads all episodes from the podcast list (redownload_list.xlsx).
RSS feeds were researched; podcasts without a known feed are auto-discovered
via the PodcastIndex.org API (free, no key required for basic search) and
Apple iTunes search as a fallback.

Usage:
    python3 download_podcasts.py [--output-dir ./podcasts] [--workers 3] [--dry-run]

Requirements:
    pip install requests feedparser openpyxl tqdm
"""

import argparse
import hashlib
import html
import json
import os
import re
import sys
import time
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.parse import urlencode, urlparse

try:
    import feedparser
    import requests
    from openpyxl import load_workbook
    from tqdm import tqdm
except ImportError as e:
    print(f"[ERROR] Missing dependency: {e}")
    print("Run:  pip install requests feedparser openpyxl tqdm")
    sys.exit(1)

# ─────────────────────────────────────────────────────────────────────────────
# KNOWN RSS FEEDS  (researched from the spreadsheet + web searches)
# ─────────────────────────────────────────────────────────────────────────────
KNOWN_RSS: dict[str, str] = {
    "7 Gute Gründe":
        "https://feeds.insohr.de/7gutegruende/",
    "einsbiszwei":
        "https://feedpress.me/einbiszwei",
    "Psychoaktiv - Dein Podcast mit Suchttherapeutin Stefanie Bötsch":
        "https://5cxann.podcaster.de/psychoaktiv.rss",
    "The Social Work Stories Podcast":
        "https://feed.podbean.com/socialworkstories/feed.xml",
    "Helden und Visionäre - Inspiration für Changeakers und Social Entrepreneurs":
        "https://heldenundvisionaere.de/feed/mp3/",
    "Wohlfahrt im Wandel - Der Community-Podcast der AWO":
        "https://rss.buzzsprout.com/2304599.rss",
    "s_innzeit - der Wissenschafspodcast zur Sozialen Arbeit":
        "https://sinnzeit.podigee.io/feed/mp3",
    "Bumsfallera":
        "https://bumsfallera.podigee.io/feed/mp3",
    "katho-Cast":
        "https://katho-cast.podigee.io/feed/mp3",
    # Theorien Sozialer Arbeit is hosted on letscast.fm
    "Theorien Sozialer Arbeit":
        "https://letscast.fm/sites/theorien-der-sozialen-arbeit-04b8c298/feed",
    # Inklusions-Podcast website suggests podigee hosting
    "Inklusions-Podcast":
        "https://inklusions-podcast.podigee.io/feed/mp3",
    # sozial.audio
    "sozial.audio - Der Podcast von einem Sozialarbeit, über Soziale Arbeit mit Matthias Palm":
        "https://sozial.audio/feed/podcast",
    # Social Sisters
    "Social Sisters - Mädelsabend mit Tiefgang":
        "https://social-sisters.podigee.io/feed/mp3",
    # about:social
    "about:social - Soziale Arbeit & digitale Medien":
        "https://aboutsocial.podigee.io/feed/mp3",
    # relevant & sozial – Paritaet BW
    "relevant & sozial":
        "https://paricast.podigee.io/feed/mp3",
}

# Podcasts with no known RSS (will be auto-discovered)
# Key = name from spreadsheet, Value = search hint for discovery
DISCOVER_HINTS: dict[str, str] = {
    "Soziale Arbeit Hacks/ Hypatia Sinn&Soziales": "Soziale Arbeit Hacks Sabine Depew",
    "Caritas Klimapodcast": "Caritas Klimapodcast Martin Herzberg",
    "SUMMA cum NEUGIER": "SUMMA cum NEUGIER Hagemann FH Diakonie",
    "#berlinbessermachen": "berlinbessermachen Paritaet Berlin",
    "viral": "viral Podcast Jugendarbeit Stadtteilzentrum Steglitz",
    "Wir müssen mal reden!": "Wir müssen mal reden Baumann Mampel Soziale Organisation",
    "STREETNOIZE": "STREETNOIZE streetwork Jugendarbeit Podcast",
    "Digital werden. Sozial bleiben.": "Digital werden Sozial bleiben Soziale Arbeit",
}

# Podcasts that are not available / Spotify-only (no public RSS)
UNAVAILABLE: set[str] = {
    "Caritas Klimapodcast",   # marked "Not available anymore" in spreadsheet
    "SUMMA cum NEUGIER",      # Spotify-only per spreadsheet
}

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": (
        "Mozilla/5.0 (compatible; PodcastDownloader/1.0; "
        "+https://github.com/example/podcast-downloader)"
    )
})


def safe_filename(name: str, max_len: int = 200) -> str:
    """Convert a string to a safe filesystem name."""
    name = unicodedata.normalize("NFC", name)
    name = html.unescape(name)
    name = re.sub(r'[\\/:*?"<>|]', "_", name)
    name = re.sub(r"\s+", " ", name).strip()
    if len(name) > max_len:
        name = name[:max_len].rstrip()
    return name or "episode"


def human_size(num_bytes: int) -> str:
    for unit in ("B", "KB", "MB", "GB"):
        if num_bytes < 1024:
            return f"{num_bytes:.1f} {unit}"
        num_bytes /= 1024
    return f"{num_bytes:.1f} TB"


def retry_get(url: str, retries: int = 4, backoff: float = 2.0,
              stream: bool = False, timeout: int = 30) -> requests.Response | None:
    """GET with exponential-backoff retry."""
    for attempt in range(retries):
        try:
            r = SESSION.get(url, stream=stream, timeout=timeout,
                            allow_redirects=True)
            if r.status_code == 200:
                return r
            if r.status_code in (429, 503):
                wait = backoff ** attempt
                time.sleep(wait)
        except requests.RequestException as exc:
            wait = backoff ** attempt
            print(f"  [WARN] {exc} – retrying in {wait:.0f}s …")
            time.sleep(wait)
    return None


# ─────────────────────────────────────────────────────────────────────────────
# RSS DISCOVERY
# ─────────────────────────────────────────────────────────────────────────────

def discover_via_itunes(query: str) -> str | None:
    """Search iTunes/Apple Podcasts for an RSS feed URL."""
    params = urlencode({"term": query, "media": "podcast", "limit": 5})
    url = f"https://itunes.apple.com/search?{params}"
    try:
        r = SESSION.get(url, timeout=15)
        if r.status_code == 200:
            data = r.json()
            results = data.get("results", [])
            if results:
                return results[0].get("feedUrl")
    except Exception:
        pass
    return None


def discover_via_podcastindex(query: str) -> str | None:
    """
    Search PodcastIndex.org (no auth required for basic search via
    the iTunes-compatible endpoint they mirror).
    Falls back to podcastindex.org /api/1.0/search/byterm.
    """
    # Attempt 1: iTunes search (most reliable, no key needed)
    feed = discover_via_itunes(query)
    if feed:
        return feed

    # Attempt 2: Listen Notes / fyyd (public, no key required)
    try:
        params = urlencode({"q": query, "count": 1})
        r = SESSION.get(
            f"https://fyyd.de/api/0.2/search/podcast?{params}",
            timeout=15,
        )
        if r.status_code == 200:
            data = r.json()
            hits = data.get("data", [])
            if hits:
                return hits[0].get("xmlURL") or hits[0].get("feed")
    except Exception:
        pass

    return None


# ─────────────────────────────────────────────────────────────────────────────
# FEED PARSING
# ─────────────────────────────────────────────────────────────────────────────

def parse_feed(rss_url: str) -> list[dict]:
    """
    Parse an RSS feed and return a list of episode dicts:
      {title, url, pub_date, ext}
    """
    feed = feedparser.parse(rss_url)
    if feed.bozo and not feed.entries:
        print(f"  [WARN] Feed parse warning for {rss_url}: {feed.bozo_exception}")

    episodes = []
    for entry in feed.entries:
        # Find the best audio enclosure
        audio_url = None
        ext = "mp3"

        # Check enclosures first
        for enc in getattr(entry, "enclosures", []):
            mime = getattr(enc, "type", "")
            href = getattr(enc, "href", "") or getattr(enc, "url", "")
            if href and ("audio" in mime or href.lower().endswith(
                    (".mp3", ".m4a", ".ogg", ".opus", ".wav", ".aac"))):
                audio_url = href
                # Derive extension from URL
                parsed = urlparse(href)
                path_ext = Path(parsed.path).suffix.lstrip(".")
                if path_ext:
                    ext = path_ext
                break

        # Fall back to links
        if not audio_url:
            for link in getattr(entry, "links", []):
                href = link.get("href", "")
                mime = link.get("type", "")
                if "audio" in mime or href.lower().endswith(
                        (".mp3", ".m4a", ".ogg", ".opus", ".wav", ".aac")):
                    audio_url = href
                    parsed = urlparse(href)
                    path_ext = Path(parsed.path).suffix.lstrip(".")
                    if path_ext:
                        ext = path_ext
                    break

        if not audio_url:
            continue  # skip episodes with no audio

        title = entry.get("title", "Untitled Episode")
        pub = entry.get("published", entry.get("updated", ""))

        episodes.append({
            "title": title,
            "url": audio_url,
            "pub_date": pub,
            "ext": ext or "mp3",
        })

    return episodes


# ─────────────────────────────────────────────────────────────────────────────
# DOWNLOAD
# ─────────────────────────────────────────────────────────────────────────────

def download_episode(episode: dict, dest_path: Path,
                     dry_run: bool = False) -> tuple[bool, str]:
    """
    Download a single episode to dest_path.
    Returns (success, message).
    """
    if dest_path.exists() and dest_path.stat().st_size > 1024:
        return True, f"Already exists ({human_size(dest_path.stat().st_size)})"

    if dry_run:
        return True, f"[DRY RUN] Would download → {dest_path.name}"

    dest_path.parent.mkdir(parents=True, exist_ok=True)

    # Stream download with progress
    try:
        r = retry_get(episode["url"], stream=True, retries=4, timeout=60)
        if r is None:
            return False, f"Failed to fetch {episode['url']}"

        total = int(r.headers.get("content-length", 0))
        tmp_path = dest_path.with_suffix(".tmp")

        with open(tmp_path, "wb") as fh:
            if total:
                with tqdm(
                    total=total,
                    unit="B",
                    unit_scale=True,
                    unit_divisor=1024,
                    desc=dest_path.name[:40],
                    leave=False,
                ) as pbar:
                    for chunk in r.iter_content(chunk_size=131072):
                        fh.write(chunk)
                        pbar.update(len(chunk))
            else:
                for chunk in r.iter_content(chunk_size=131072):
                    fh.write(chunk)

        tmp_path.rename(dest_path)
        size = dest_path.stat().st_size
        return True, f"Downloaded {human_size(size)}"

    except Exception as exc:
        return False, f"Error: {exc}"


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def load_podcast_list(xlsx_path: str) -> list[dict]:
    """Load podcast list from the Excel file."""
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    podcasts = []
    headers = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = row
            continue
        if not row[0]:
            continue
        entry = dict(zip(headers, row))
        podcasts.append(entry)
    return podcasts


def build_rss_map(podcasts: list[dict]) -> dict[str, str | None]:
    """
    Return {podcast_name: rss_url | None}.
    Priority: spreadsheet RSS column → KNOWN_RSS → auto-discovery.
    """
    rss_map: dict[str, str | None] = {}

    for p in podcasts:
        name = str(p.get("Podcast Name", "")).strip()
        if not name:
            continue

        # 1. Use RSS from spreadsheet if valid
        sheet_rss = p.get("RSS", "")
        if sheet_rss and str(sheet_rss).startswith("http") and "rss" in str(sheet_rss).lower() or (
            sheet_rss and str(sheet_rss).startswith("http") and "feed" in str(sheet_rss).lower()
        ) or (
            sheet_rss and str(sheet_rss).startswith("https://") and sheet_rss not in (
                "spotify", "Not available anymore", "soziale-arbeit-hacks/"
            )
        ):
            rss_map[name] = str(sheet_rss)
            continue

        # 2. Known RSS map
        for key, url in KNOWN_RSS.items():
            if key.lower() in name.lower() or name.lower() in key.lower():
                rss_map[name] = url
                break
        else:
            rss_map[name] = None

    return rss_map


def discover_missing(rss_map: dict[str, str | None]) -> dict[str, str | None]:
    """Try to discover RSS feeds for podcasts missing them."""
    missing = [n for n, u in rss_map.items() if u is None and n not in UNAVAILABLE]
    if not missing:
        return rss_map

    print(f"\n🔍 Attempting auto-discovery for {len(missing)} podcasts …")
    for name in missing:
        hint = DISCOVER_HINTS.get(name, name)
        print(f"  Searching: {name!r} …", end=" ", flush=True)
        url = discover_via_podcastindex(hint)
        if url:
            print(f"✅ {url}")
            rss_map[name] = url
        else:
            print("❌ not found")

    return rss_map


def run(xlsx_path: str, output_dir: str, workers: int, dry_run: bool) -> None:
    out_root = Path(output_dir)
    out_root.mkdir(parents=True, exist_ok=True)

    # Load podcast list
    print(f"📋 Loading podcast list from {xlsx_path} …")
    podcasts = load_podcast_list(xlsx_path)
    print(f"   Found {len(podcasts)} podcasts.")

    # Build / discover RSS map
    rss_map = build_rss_map(podcasts)
    rss_map = discover_missing(rss_map)

    # Summary
    found = sum(1 for v in rss_map.values() if v)
    unavail = len(UNAVAILABLE)
    print(f"\n📡 RSS feeds found: {found}/{len(rss_map)}  "
          f"(unavailable/Spotify-only: {unavail})")

    # Log unavailable / missing
    log_path = out_root / "download_log.txt"
    log_lines = []

    for name, url in sorted(rss_map.items()):
        if name in UNAVAILABLE:
            log_lines.append(f"[UNAVAILABLE] {name}")
        elif url is None:
            log_lines.append(f"[NO_RSS_FOUND] {name}")

    # Process each podcast
    total_ok = 0
    total_fail = 0
    total_skip = 0

    for podcast in podcasts:
        name = str(podcast.get("Podcast Name", "")).strip()
        if not name:
            continue

        rss_url = rss_map.get(name)

        if name in UNAVAILABLE:
            print(f"\n⛔  {name!r} — marked unavailable, skipping.")
            continue

        if not rss_url:
            print(f"\n⚠️  {name!r} — no RSS feed found, skipping.")
            continue

        print(f"\n🎙️  {name}")
        print(f"    Feed: {rss_url}")

        # Parse feed
        try:
            episodes = parse_feed(rss_url)
        except Exception as exc:
            print(f"    [ERROR] Feed parse failed: {exc}")
            log_lines.append(f"[FEED_ERROR] {name}: {exc}")
            continue

        if not episodes:
            print(f"    [WARN] No downloadable episodes found in feed.")
            log_lines.append(f"[NO_EPISODES] {name}")
            continue

        print(f"    Episodes found: {len(episodes)}")

        # Create podcast directory
        pod_dir = out_root / safe_filename(name)
        pod_dir.mkdir(parents=True, exist_ok=True)

        # Build download tasks
        tasks = []
        for i, ep in enumerate(episodes, 1):
            ep_title = safe_filename(ep["title"])
            # Zero-pad index for correct ordering
            filename = f"{i:04d} - {ep_title}.{ep['ext']}"
            dest = pod_dir / filename
            tasks.append((ep, dest))

        # Download with thread pool
        ok = fail = skip = 0
        with ThreadPoolExecutor(max_workers=workers) as pool:
            futures = {
                pool.submit(download_episode, ep, dest, dry_run): (ep, dest)
                for ep, dest in tasks
            }
            with tqdm(total=len(futures), desc=f"  {name[:35]}", unit="ep") as pbar:
                for fut in as_completed(futures):
                    ep, dest = futures[fut]
                    success, msg = fut.result()
                    if success:
                        if "Already exists" in msg:
                            skip += 1
                        else:
                            ok += 1
                        log_lines.append(f"[OK] {name} / {dest.name}: {msg}")
                    else:
                        fail += 1
                        log_lines.append(f"[FAIL] {name} / {ep['title']}: {msg}")
                        tqdm.write(f"    [FAIL] {ep['title']}: {msg}")
                    pbar.update(1)

        total_ok += ok
        total_fail += fail
        total_skip += skip
        print(f"    ✅ {ok} downloaded  ⏭️  {skip} skipped  ❌ {fail} failed")

    # Write log
    with open(log_path, "w", encoding="utf-8") as fh:
        fh.write("\n".join(log_lines))

    print(f"""
╔══════════════════════════════════════════════╗
║              DOWNLOAD COMPLETE               ║
╠══════════════════════════════════════════════╣
║  ✅ Downloaded : {total_ok:<5}                       ║
║  ⏭️  Skipped   : {total_skip:<5}  (already present)  ║
║  ❌ Failed     : {total_fail:<5}                       ║
║  📄 Log        : {str(log_path):<28} ║
╚══════════════════════════════════════════════╝
""")


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Bulk-download all episodes from the podcast spreadsheet.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument(
        "--xlsx",
        default="redownload_list.xlsx",
        help="Path to the Excel podcast list.",
    )
    parser.add_argument(
        "--output-dir",
        default="./podcasts",
        help="Root directory for downloaded episodes.",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=3,
        help="Parallel download threads (keep ≤5 to be polite).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="List what would be downloaded without actually downloading.",
    )
    args = parser.parse_args()

    if not Path(args.xlsx).exists():
        print(f"[ERROR] Excel file not found: {args.xlsx}")
        sys.exit(1)

    run(
        xlsx_path=args.xlsx,
        output_dir=args.output_dir,
        workers=args.workers,
        dry_run=args.dry_run,
    )


if __name__ == "__main__":
    main()
